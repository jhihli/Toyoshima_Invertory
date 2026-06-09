"""
Management command: send_mpn_report

Usage:
    python manage.py send_mpn_report
    python manage.py send_mpn_report --triggered-by manual

The cron job (* * * * *) runs every minute; this command checks whether the
current local time matches config.send_time (HH:MM) and skips if already
sent today. The 'Send Now' API endpoint calls it with --triggered-by manual,
which bypasses both the time check and the auto_send_enabled toggle.
"""
import io
import traceback

from django.conf import settings
from django.core.mail import EmailMessage
from django.core.management.base import BaseCommand
from django.db.models import Max
from django.utils import timezone

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from product.models import MPN, MPNReportConfig, MPNReportEmail


class Command(BaseCommand):
    help = 'Generate MPN report Excel and email it to the configured recipient.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--triggered-by',
            default='cron',
            choices=['cron', 'manual'],
            help='cron (respects toggle) or manual (always sends)',
        )

    def handle(self, *args, **options):
        triggered_by = options['triggered_by']
        config = MPNReportConfig.get_config()

        if triggered_by == 'cron':
            if not config.auto_send_enabled:
                self.stdout.write('Auto-send is disabled. Skipping.')
                return

            # Check if current local time matches the configured send_time (HH:MM)
            now = timezone.localtime(timezone.now())
            cfg_time = config.send_time
            if now.hour != cfg_time.hour or now.minute != cfg_time.minute:
                return  # not the right minute — silent skip

            # Guard against duplicate sends within the same day
            today = now.date()
            already_sent = MPNReportEmail.objects.filter(
                sent_at__date=today, triggered_by='cron', status='ok'
            ).exists()
            if already_sent:
                self.stdout.write('Already sent today. Skipping.')
                return

        if not config.recipient:
            self.stderr.write('No recipient configured in MPNReportConfig. Aborting.')
            return

        from_email = settings.DEFAULT_FROM_EMAIL or settings.EMAIL_HOST_USER
        if not from_email:
            err = 'EMAIL_HOST_USER / DEFAULT_FROM_EMAIL is not set in .env. Configure SMTP settings before sending.'
            self.stderr.write(err)
            MPNReportEmail.objects.create(
                sent_to=config.recipient, cc=config.cc,
                status='error', error=err, triggered_by=triggered_by,
            )
            return

        self.stdout.write(f'Generating MPN report (triggered_by={triggered_by})…')

        try:
            excel_bytes = self._build_excel()
            self._send_email(excel_bytes, config.recipient, config.cc, from_email)
            MPNReportEmail.objects.create(
                sent_to=config.recipient,
                cc=config.cc,
                status='ok',
                triggered_by=triggered_by,
            )
            self.stdout.write(self.style.SUCCESS(f'Report sent to {config.recipient}'))
        except Exception:
            err_text = traceback.format_exc()
            MPNReportEmail.objects.create(
                sent_to=config.recipient,
                cc=config.cc,
                status='error',
                error=err_text,
                triggered_by=triggered_by,
            )
            self.stderr.write(f'Failed to send report:\n{err_text}')
            raise  # non-zero exit so cron logs the failure

    # ─── Excel builder ─────────────────────────────────────────────────────────

    def _build_excel(self) -> bytes:
        """
        Mirrors the frontend handleExport columns exactly:
          MPN | Date Processed | Chip MPN | Chips Failed |
          Failure Rate | BOARDS (Scanned) | CUTBOARD COST | CHIP COST
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'MPN Export'

        headers = [
            'MPN',
            'Date Processed',
            'Chip MPN',
            'Chips Failed',
            'Failure Rate (BOARDS Scanned / Chips Failed)',
            'BOARDS (Scanned)',
            'CUTBOARD COST',
            'CHIP COST',
        ]
        col_widths = [22, 15, 24, 13, 44, 17, 15, 13]

        header_fill = PatternFill('solid', fgColor='D9E1F2')
        header_font = Font(bold=True, size=11)
        left = Alignment(horizontal='left')

        for c_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=c_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = left
            ws.column_dimensions[get_column_letter(c_idx)].width = width

        ws.freeze_panes = 'A2'

        row_num = 2
        mpns = MPN.objects.prefetch_related('chips__brand', 'boards').order_by('name')

        for mpn in mpns:
            chips = list(mpn.chips.all())
            board_count = mpn.boards.count()

            latest_result = mpn.boards.aggregate(latest=Max('scanned_at'))['latest']
            latest_date = latest_result.strftime('%Y-%m-%d') if latest_result else ''

            chip_type_count = len(chips)
            chip_cost = (
                round(float(mpn.cutboard_cost) / chip_type_count, 4)
                if mpn.cutboard_cost and chip_type_count > 0
                else None
            )

            if not chips:
                row_data = [
                    mpn.name, latest_date, '', '',
                    None,
                    board_count,
                    float(mpn.cutboard_cost) if mpn.cutboard_cost is not None else '',
                    '',
                ]
                for c_idx, val in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_num, column=c_idx, value=val)
                    cell.alignment = left
                row_num += 1
            else:
                for chip in chips:
                    failure_rate = (
                        chip.cut_fail / board_count
                        if chip.cut_fail is not None and chip.cut_fail > 0 and board_count > 0
                        else 0
                    )
                    row_data = [
                        mpn.name,
                        latest_date,
                        chip.chip_mpn or '',
                        chip.cut_fail if chip.cut_fail is not None else 0,
                        failure_rate,
                        board_count,
                        float(mpn.cutboard_cost) if mpn.cutboard_cost is not None else '',
                        chip_cost if chip_cost is not None else '',
                    ]
                    for c_idx, val in enumerate(row_data, start=1):
                        cell = ws.cell(row=row_num, column=c_idx, value=val)
                        cell.alignment = left
                        if c_idx == 5 and val is not None:
                            cell.number_format = '0.00%'
                    row_num += 1

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ─── Email sender ──────────────────────────────────────────────────────────

    def _send_email(self, excel_bytes: bytes, recipient: str, cc: str, from_email: str) -> None:
        today = timezone.localdate().strftime('%Y-%m-%d')
        subject = f'MPN Report — {today}'
        body = (
            'Hi Team,\n\n'
            f'Please find the attached MPN detail report for {today}.\n'
            'This email was sent automatically by the Toyoshima Inventory system.\n\n\n\n'
            'Best regards,\n'
            'Toyoshima Green Tech\n'
        )
        to_list = [e.strip() for e in recipient.split(',') if e.strip()]
        cc_list = [e.strip() for e in cc.split(',') if e.strip()] if cc else []
        msg = EmailMessage(
            subject=subject,
            body=body,
            from_email=from_email,
            to=to_list,
            cc=cc_list,
        )
        filename = f'mpn_report_{today}.xlsx'
        msg.attach(
            filename,
            excel_bytes,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        msg.send(fail_silently=False)
